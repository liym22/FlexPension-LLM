#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Evaluate the DKI teacher predictions for all CHFS 2019 samples (seed 42)."""

import os
import json
import sys
import pandas as pd
import numpy as np
from datetime import datetime
from sklearn.metrics import f1_score, precision_score, recall_score, confusion_matrix
import warnings

warnings.filterwarnings("ignore")

# ==================== Path configuration ====================
from pathlib import Path

_CODE_DIR = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(_CODE_DIR))

from config.paths import (
    CHFS2019_DKI_GROUND_TRUTH_FILE,
    TEACHER_INFERENCE_CHFS2019_CLAUDE_EVALUATION_DIR,
    TEACHER_INFERENCE_CHFS2019_CLAUDE_RESULTS_DIR,
)

GROUND_TRUTH_PATH = str(CHFS2019_DKI_GROUND_TRUTH_FILE)
JSON_INPUT_DIR = str(TEACHER_INFERENCE_CHFS2019_CLAUDE_RESULTS_DIR)
EVAL_OUTPUT_DIR = str(TEACHER_INFERENCE_CHFS2019_CLAUDE_EVALUATION_DIR)
os.makedirs(EVAL_OUTPUT_DIR, exist_ok=True)

MODEL_SHORT_NAME = "claude45sonnet"
SEEDS = [42]
TEMPERATURE = 0.5

# ==================== Load ground truth ====================
with open(GROUND_TRUTH_PATH, "r", encoding="utf-8") as f:
    ground_truths = json.load(f)

gt_dict = {
    f"{item['household_id']}-{item['individual_id']}": {
        "decision": item.get("decision", ""),
        "type": item.get("type", ""),
    }
    for item in ground_truths
}
print(f"✓ Loaded ground truth: {len(gt_dict)} samples")


# ==================== Load inference results ====================
def load_results():
    results = {}
    for seed in SEEDS:
        fname = f"{MODEL_SHORT_NAME}_seed{seed}_temp{str(TEMPERATURE).replace('.', '')}_results.json"
        fpath = os.path.join(JSON_INPUT_DIR, fname)
        if not os.path.exists(fpath):
            print(f"⚠ Result file not found: {fname}")
            continue
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            results[seed] = data
            total = len(data["tests"])
            api_ok = sum(1 for t in data["tests"] if t.get("success"))
            parse_ok = sum(1 for t in data["tests"] if t.get("parse_success"))
            print(f"✓ Loaded seed {seed}: {total} samples | API✓ {api_ok} | JSON✓ {parse_ok}")
        except Exception as e:
            print(f"⚠ Skipping invalid file {fname}: {str(e)[:60]}")
    return results


# ==================== Evaluate one seed ====================
def evaluate_seed(seed_data, gt_dict):
    y_true_action, y_pred_action = [], []
    y_true_type, y_pred_type = [], []

    rows = []  # Store row-level details for the Excel report.

    for test in seed_data["tests"]:
        sample_id = test["sample_id"]
        gt = gt_dict.get(sample_id, {})
        decision = gt.get("decision", "")
        gt_type = gt.get("type", "")

        pred_action = test.get("predicted_action", "N/A")
        pred_type = test.get("predicted_insurance_type", "N/A")
        parse_ok = test.get("parse_success", False)

        act_true = 1 if decision == "参保" else 0
        act_pred = 1 if pred_action == "参保" else 0
        y_true_action.append(act_true)
        y_pred_action.append(act_pred)

        action_correct = int(pred_action == decision) if parse_ok else None
        type_correct = None

        if parse_ok and pred_action == "参保" and decision == "参保":
            true_lbl = 1 if gt_type == "城镇职工养老保险" else 0
            pred_lbl = 1 if pred_type == "城镇职工养老保险" else 0
            y_true_type.append(true_lbl)
            y_pred_type.append(pred_lbl)
            type_correct = int(pred_type == gt_type)

        rows.append(
            {
                "sample_id": sample_id,
                "household_id": test.get("household_id", ""),
                "individual_id": test.get("individual_id", ""),
                "api_success": test.get("success", False),
                "parse_success": parse_ok,
                "gt_decision": decision,
                "gt_type": gt_type,
                "pred_action": pred_action if parse_ok else "N/A",
                "pred_type": pred_type if parse_ok else "N/A",
                "pred_annual_payment": test.get("predicted_annual_payment", ""),
                "pred_main_reason": test.get("predicted_main_reason", ""),
                "action_correct": action_correct,
                "type_correct": type_correct,
            }
        )

    # Participation-action metrics
    try:
        action_f1 = f1_score(y_true_action, y_pred_action, zero_division=0)
        action_prec = precision_score(y_true_action, y_pred_action, zero_division=0)
        action_recall = recall_score(y_true_action, y_pred_action, zero_division=0)
        action_acc = (
            sum(a == b for a, b in zip(y_true_action, y_pred_action))
            / len(y_true_action)
            if y_true_action
            else 0
        )
        cm = confusion_matrix(y_true_action, y_pred_action, labels=[0, 1]).tolist()
    except Exception:
        action_f1 = action_prec = action_recall = action_acc = 0.0
        cm = []

    # Pension-type metrics
    if y_true_type:
        try:
            type_f1 = f1_score(y_true_type, y_pred_type, zero_division=0)
            type_prec = precision_score(y_true_type, y_pred_type, zero_division=0)
            type_recall = recall_score(y_true_type, y_pred_type, zero_division=0)
            type_acc = sum(a == b for a, b in zip(y_true_type, y_pred_type)) / len(
                y_true_type
            )
        except Exception:
            type_f1 = type_prec = type_recall = type_acc = 0.0
    else:
        type_f1 = type_prec = type_recall = type_acc = None

    return {
        "action_metrics": {
            "f1": action_f1,
            "precision": action_prec,
            "recall": action_recall,
            "accuracy": action_acc,
            "confusion_matrix": cm,
        },
        "type_metrics": {
            "f1": type_f1,
            "precision": type_prec,
            "recall": type_recall,
            "accuracy": type_acc,
        },
        "n_total": len(y_true_action),
        "n_insured_true": int(sum(y_true_action)),
        "n_insured_pred": int(sum(y_pred_action)),
        "n_type_eval": len(y_true_type),
        "rows": rows,
    }


# ==================== Main evaluation workflow ====================
def main():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    print("=" * 80)
    print(f"Full-Sample Evaluation - {MODEL_SHORT_NAME}  Seeds: {SEEDS}")
    print("=" * 80)

    all_results = load_results()
    if not all_results:
        print("❌ No valid result files found; exiting")
        return

    # Evaluate each seed; the reported teacher run uses seed 42.
    seed_evals = {}
    for seed, data in all_results.items():
        print(f"\n▶ Evaluating seed {seed}...")
        ev = evaluate_seed(data, gt_dict)
        seed_evals[seed] = ev

        am = ev["action_metrics"]
        tm = ev["type_metrics"]
        print(
            f"  • Total: {ev['n_total']} | Actual participants: {ev['n_insured_true']} | Predicted participants: {ev['n_insured_pred']}"
        )
        print(
            f"  • Action  — Acc: {am['accuracy']:.4f} | F1: {am['f1']:.4f} | "
            f"Prec: {am['precision']:.4f} | Recall: {am['recall']:.4f}"
        )
        if tm["f1"] is not None:
            print(
                f"  • Type    — Acc: {tm['accuracy']:.4f} | F1: {tm['f1']:.4f} | "
                f"Prec: {tm['precision']:.4f} | Recall: {tm['recall']:.4f} "
                f"(n={ev['n_type_eval']})"
            )
        else:
            print(f"  • Type    — No evaluable samples")
        if am["confusion_matrix"]:
            print("  • Confusion matrix (rows=actual, columns=predicted, "
                  "0=non-enrollment, 1=enrollment):")
            for row in am["confusion_matrix"]:
                print(f"      {row}")

    # ==================== Save row-level Excel diagnostics ====================
    seed = SEEDS[0]
    if seed in seed_evals:
        ev = seed_evals[seed]
        df = pd.DataFrame(ev["rows"])
        excel_path = os.path.join(
            EVAL_OUTPUT_DIR, f"detail_seed{seed}_{timestamp}.xlsx"
        )
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Details")
            ws = writer.sheets["Details"]
            # Freeze the header row.
            from openpyxl.styles import PatternFill, Font

            ws.freeze_panes = "A2"
            # Highlight incorrect predictions.
            red_fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )
            green_fill = PatternFill(
                start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
            )
            action_col = df.columns.get_loc("action_correct") + 1
            for row_idx, val in enumerate(df["action_correct"], start=2):
                if val == 0:
                    ws.cell(row=row_idx, column=action_col).fill = red_fill
                elif val == 1:
                    ws.cell(row=row_idx, column=action_col).fill = green_fill
        print(f"\n✓ Detailed Excel report: {os.path.basename(excel_path)}")

    # ==================== Save aggregate metrics as JSON ====================
    # Convert NumPy scalar values to native Python values.
    def to_native(obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, dict):
            return {k: to_native(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [to_native(i) for i in obj]
        return obj

    summary = {
        "evaluation_time": datetime.now().isoformat(),
        "model": MODEL_SHORT_NAME,
        "seeds": SEEDS,
        "ground_truth_samples": len(gt_dict),
        "actual_insured": sum(
            1 for gt in gt_dict.values() if gt.get("decision") == "参保"
        ),
        "seed_evaluations": {
            str(seed): to_native(
                {
                    "action_metrics": ev["action_metrics"],
                    "type_metrics": ev["type_metrics"],
                    "n_total": ev["n_total"],
                    "n_insured_true": ev["n_insured_true"],
                    "n_insured_pred": ev["n_insured_pred"],
                    "n_type_eval": ev["n_type_eval"],
                }
            )
            for seed, ev in seed_evals.items()
        },
    }

    json_path = os.path.join(EVAL_OUTPUT_DIR, f"metrics_{timestamp}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"✓ Metrics summary JSON: {os.path.basename(json_path)}")

    # ==================== Print final summary ====================
    print("\n" + "=" * 80)
    print("Evaluation Summary")
    print("=" * 80)
    header = (
        f"{'Seed':<6} {'N':<8} {'Actual':<8} {'Pred.':<8} "
        f"{'ActAcc':<8} {'ActF1':<8} {'ActPrec':<8} {'ActRec':<8} "
        f"{'TypAcc':<8} {'TypF1':<8} {'TypPrec':<8} {'TypRec':<8}"
    )
    print(header)
    print("-" * len(header))
    for seed, ev in seed_evals.items():
        am = ev["action_metrics"]
        tm = ev["type_metrics"]
        print(
            f"{seed:<6} {ev['n_total']:<8} {ev['n_insured_true']:<8} {ev['n_insured_pred']:<8} "
            f"{am['accuracy']:<8.4f} {am['f1']:<8.4f} {am['precision']:<8.4f} {am['recall']:<8.4f} "
            f"{(tm['accuracy'] or 0):<8.4f} {(tm['f1'] or 0):<8.4f} "
            f"{(tm['precision'] or 0):<8.4f} {(tm['recall'] or 0):<8.4f}"
        )

    print(f"\n✓ Output directory: {EVAL_OUTPUT_DIR}")


if __name__ == "__main__":
    try:
        from sklearn.metrics import (
            f1_score,
            precision_score,
            recall_score,
            confusion_matrix,
        )
    except ImportError:
        print("❌ Install required packages: pip install scikit-learn pandas openpyxl")
        exit(1)
    main()
