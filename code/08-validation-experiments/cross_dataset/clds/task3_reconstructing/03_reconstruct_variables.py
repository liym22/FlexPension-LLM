#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
03_reconstruct_variables.py
Reconstruct all variables required by the prompts for the 500 cases sampled in
task 2, and write reconstructed_500.csv.

Data sources:
  - 18个人.dta (individual records; FID2018 is the household key)
  - 18家庭.dta (household income, assets, debt, children, and pensions)
  - data/raw/policy/province_policy_2018.xlsx (2018 provincial pension policy)

Key differences from CFPS/CHIP:
  - Gender codes: 1=男, 2=女 (Igender)
  - Age: 2018 - birthyear; no direct age variable
  - Flexible employment: derived from I3a_9 and I3a1_5
  - Household composition: derived from birthyear_1 through birthyear_22 in
    18家庭.dta
  - Household enrollment count: aggregated by FID2018 over all person records,
    which include respondents only
  - Household benefit-recipient count: unavailable and always "不清楚"
  - Missing-value codes: {99997, 99998, 99999} (拒绝/不适用/不清楚)
  - Risk preference: I7_14_6_w16 (不向银行贷款买车)
  - Economic expectations: I7_6_3 (经济满意度)
  - 历史参保类型: always "不清楚" because historical data are unavailable
  - 城乡: areatype (1=农村, 2=城镇)

Asset units, verified against the source data:
  - f3_4_4, f3_5_6, f3_6_6: 10,000 CNY; multiply by 10,000
  - f3_11_5_1/2/3/4:         CNY, not 10,000 CNY
  - f5f5_2b:                  CNY
  - f4_10_3, f4_10_3_panel:  CNY
  - f4_11_2_panel through f4_14_2: CNY
  - f4_16_1:                  CNY
  - I3a2_4_total:             CNY, not 10,000 CNY
  - I3a3_23:                  10,000 CNY; multiply by 10,000
"""

import math
import sys
from pathlib import Path
import numpy as np
import pandas as pd
import pyreadstat
import openpyxl

# Paths
SCRIPT_DIR  = Path(__file__).resolve().parent
TASK1_DIR   = SCRIPT_DIR.parent / "task1"
TASK2_DIR   = SCRIPT_DIR.parent / "task2_sampling"
PROJECT_ROOT = Path(__file__).resolve().parents[5]
sys.path.insert(0, str(PROJECT_ROOT / "code"))
from common.household_enrollment import count_other_enrolled_members

CLDS_DIR    = PROJECT_ROOT / "data" / "raw" / "clds2018"
PERSON_DTA  = CLDS_DIR / "18个人.dta"
FAM_DTA     = CLDS_DIR / "18家庭.dta"
POLICY_XLS  = PROJECT_ROOT / "data" / "raw" / "policy" / "province_policy_2018.xlsx"
OUT_CSV     = SCRIPT_DIR / "reconstructed_500.csv"
LOG_PATH    = SCRIPT_DIR / "03_reconstruct_log.txt"

SURVEY_YEAR = 2018
MISSING_VALS = {99997, 99998, 99999}

# Logging
_log = open(LOG_PATH, "w", encoding="utf-8")

def log(msg=""):
    print(msg)
    _log.write(msg + "\n")
    _log.flush()

log("=" * 70)
log("Task 3  CLDS 2018 variable reconstruction")
log("=" * 70)

# Static mappings
PROV_CODE_TO_NAME = {
    11: "北京市",   12: "天津市",   13: "河北省",   14: "山西省",
    15: "内蒙古自治区", 21: "辽宁省", 22: "吉林省",   23: "黑龙江省",
    31: "上海市",   32: "江苏省",   33: "浙江省",   34: "安徽省",
    35: "福建省",   36: "江西省",   37: "山东省",   41: "河南省",
    42: "湖北省",   43: "湖南省",   44: "广东省",   45: "广西壮族自治区",
    46: "海南省",   50: "重庆市",   51: "四川省",   52: "贵州省",
    53: "云南省",   54: "西藏自治区", 61: "陕西省", 62: "甘肃省",
    63: "青海省",   64: "宁夏回族自治区", 65: "新疆维吾尔自治区",
}
REGION_MAP = {
    11: "东部", 12: "东部", 13: "东部", 31: "东部", 32: "东部",
    33: "东部", 35: "东部", 37: "东部", 44: "东部", 46: "东部",
    14: "中部", 34: "中部", 36: "中部", 41: "中部", 42: "中部", 43: "中部",
    15: "西部", 45: "西部", 50: "西部", 51: "西部", 52: "西部",
    53: "西部", 54: "西部", 61: "西部", 62: "西部", 63: "西部",
    64: "西部", 65: "西部",
    21: "东北", 22: "东北", 23: "东北",
}
EDU_MAP = {
    1: "未上过学", 2: "小学", 3: "初中", 4: "普通高中",
    5: "职业高中", 6: "技校", 7: "中专", 8: "大专",
    9: "本科", 10: "硕士", 11: "博士",
}
HEALTH_MAP = {
    1: "非常健康", 2: "健康", 3: "一般", 4: "比较不健康", 5: "非常不健康",
}
HUKOU_MAP = {
    1: "农业户口", 2: "非农户口",
    3: "居民户口（原农业）", 4: "居民户口（原非农）", 99: "其他",
}
IND_MAP = {
    1:  "农、林、牧、渔业",
    2:  "采掘业",
    3:  "制造业",
    4:  "电力、煤气及水的生产和供给业",
    5:  "建筑业",
    6:  "地质勘查业、水利管理业",
    7:  "交通运输、仓储及邮电通信业",
    8:  "批发和零售贸易、餐饮业",
    9:  "金融保险业",
    10: "房地产业",
    11: "社会服务业",
    12: "卫生、体育和社会福利业",
    13: "教育、文化艺术和广播电影电视业",
    14: "科学研究和综合技术服务业",
    15: "国家机关、党政机关和社会团体",
    16: "其他行业",
}
OCC_MAP = {  # First digit of I3a_7code.
    1: "管理人员", 2: "专业技术人员", 3: "办事人员",
    4: "商业服务业人员", 5: "农林牧渔技能人员",
    6: "生产运输操作人员", 7: "军人", 8: "其他从业人员",
}
UNIT_MAP = {  # I3a_9
    1:  "党政机关、人民团体、军队",
    2:  "国有/集体事业单位",
    3:  "国营企业",
    4:  "集体企业",
    5:  "村居委会等自治组织",
    6:  "民营、私营企业",
    7:  "外资、合资企业",
    8:  "民办非企业、社团等社会组织",
    9:  "个体工商户",
    11: "自由职业者",
    12: "无固定工作者",
}
WORK_NATURE_MAP = {  # I3a_16
    1: "雇员", 2: "雇主", 3: "自雇", 4: "务农",
}
RISK_MAP = {  # Agreement indicates aversion to financing a car with bank credit.
    1: "极度厌恶风险", 2: "厌恶风险", 3: "风险中性",
    4: "偏好风险", 5: "极度偏好风险",
}
EXPECT_MAP = {  # I7_6_3 economic satisfaction.
    1: "非常悲观", 2: "悲观", 3: "一般", 4: "乐观", 5: "非常乐观",
}

def is_missing(v):
    if v is None: return True
    if isinstance(v, float) and math.isnan(v): return True
    try:
        return int(v) in MISSING_VALS
    except (ValueError, TypeError):
        return False

def safe_val(v, default=None):
    return default if is_missing(v) else v

def safe_float(v, default=0.0):
    if is_missing(v): return default
    try:
        f = float(v)
        return default if math.isnan(f) else f
    except (ValueError, TypeError):
        return default

def fmt_yuan(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return "不清楚"
    v = float(v)
    if v == 0:
        return "0元"
    if abs(v) >= 1e8:
        return f"{v/1e8:.1f}亿元"
    if abs(v) >= 1e4:
        return f"{v/1e4:.0f}万元"
    return f"{v:.0f}元"

# Load sampled cases
log("\n[1] Loading sampled_500.csv and filtered_flexible_workers.csv...")
sampled = pd.read_csv(TASK2_DIR / "sampled_500.csv")
filtered = pd.read_csv(TASK1_DIR / "filtered_flexible_workers.csv")
log(f"  Sampled records: {len(sampled)}")
log(f"  Candidate-pool records: {len(filtered)}")

# IID2018 and FID2018 exceed exact float64 integer precision.
# Use the same float64 representation from DTA and CSV as the matching key.
def to_float_key(series):
    """
    Convert IID/FID values to Int64 keys for matching.

    CLDS IDs are 15-16 digit integers that exceed float64 precision. After a
    CSV round trip, their string form may include a decimal (for example,
    "…0001.5"); rounding before conversion recovers the original integer ID.
    """
    def _conv(v):
        try:
            return int(round(float(v)))
        except (ValueError, TypeError):
            return pd.NA
    return series.apply(_conv).astype("Int64")

sampled["_iid_f"] = to_float_key(sampled["IID2018"])
sampled["_fid_f"] = to_float_key(sampled["FID2018"])
filtered["_iid_f"] = to_float_key(filtered["IID2018"])

sampled_iids_float  = set(sampled["_iid_f"].dropna())
filtered_iids_float = set(filtered["_iid_f"].dropna())
sampled_fids_float  = set(sampled["_fid_f"].dropna())

# Load person table
log("\n[2] Loading individual records from 18个人.dta...")
PERSON_COLS = [
    "IID2018", "FID2018",
    "Igender", "birthyear", "areatype",
    "PROV2018", "I1_3_1_psu", "I1_14",
    "I2_1", "I9_4_1",
    "I3a_9", "I3a1_5", "I3a_16", "I3a_8", "I3a_7code",
    "I3a_6",
    "I3a2_4_total", "I3a3_23",
    "I7_14_6_w16", "I7_6_3",
    # Pension fields used for household enrollment counts.
    "I1_20_2", "I1_20_3", "I1_20_4", "I1_20_50",
]
ind_full, meta = pyreadstat.read_dta(str(PERSON_DTA), usecols=PERSON_COLS)
ind_full["_iid_f"] = to_float_key(ind_full["IID2018"])
ind_full["_fid_f"] = to_float_key(ind_full["FID2018"])
log(f"  Loaded: {len(ind_full):,} rows")

# Restrict to the 500 sampled float64 IDs.
ind = ind_full[ind_full["_iid_f"].isin(sampled_iids_float)].copy()
log(f"  Matched sampled cases by float key: {len(ind)}")

# Load household table
log("\n[3] Loading household records from 18家庭.dta...")
FAM_COLS_BASE = [
    "FID2018", "number",
    "f4_1", "f4_25", "f4_8", "f4_17_1",
    # Residential property (10,000 CNY)
    "f3_4_4", "f3_5_6", "f3_6_6",
    # Other housing assets (CNY)
    "f3_11_5_1", "f3_11_5_2", "f3_11_5_3", "f3_11_5_4",
    # Property in the home village (CNY)
    "f5f5_2b",
    # Vehicles (CNY)
    "f4_10_3", "f4_10_3_panel", "f4_11_2_panel", "f4_12_2_panel",
    "f4_13_2_panel", "f4_14_2",
    # Accounts receivable (CNY)
    "f4_16_1",
    # Children and older adults, derived from birth year
] + [f"birthyear_{i}" for i in range(1, 23)]

fam_full, _ = pyreadstat.read_dta(str(FAM_DTA), usecols=FAM_COLS_BASE)
fam_full["_fid_f"] = to_float_key(fam_full["FID2018"])
fam = fam_full[fam_full["_fid_f"].isin(sampled_fids_float)].copy()
# Index household rows by float64 household ID.
fam = fam.set_index("_fid_f")
log(f"  Loaded full household table: {len(fam_full):,} rows")
log(f"  Matched sampled households by float key: {len(fam)} rows")

# Compute household enrollment counts
log("\n[4] Computing household enrollment counts by FID2018...")
def is_insured(row):
    return any(row.get(f"I1_20_{c}", 2) == 1
               for c in ["2", "3", "4", "50"])

ind_full["_insured"] = (
    (ind_full["I1_20_2"] == 1) |
    (ind_full["I1_20_3"] == 1) |
    (ind_full["I1_20_4"] == 1) |
    (ind_full["I1_20_50"] == 1)
).astype(int)

sampled["家庭参保人数"] = count_other_enrolled_members(
    ind_full,
    sampled,
    household_col="_fid_f",
    person_col="_iid_f",
    enrolled_col="_insured",
)
log("  Household enrollment counts computed from respondent records")

# Load policy data
log("\n[5] Loading policy data...")
wb = openpyxl.load_workbook(str(POLICY_XLS), read_only=True, data_only=True)
ws = wb["总表"]
header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
policy_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    policy_rows.append(dict(zip(header, row)))
wb.close()

policy_by_prov = {}
for r in policy_rows:
    pname = str(r.get("省名", "")).strip()
    if pname:
        policy_by_prov[pname] = r

log(f"  Provinces in policy table: {len(policy_by_prov)}")

POLICY_COLS = [
    "缴费档次与补贴明细（年，分档次）",
    "基础养老金（年）",
    "社平工资（年）",
    "缴费指数（下限）",
    "缴费指数（上限）",
    "可选档次规则",
    "缴费基数（年，下限）",
    "缴费基数（年，上限）",
    "缴费金额（年，下限）",
    "缴费金额（年，上限）",
]
COL_RENAME = {
    "缴费档次与补贴明细（年，分档次）": "缴费档次与补贴明细",
    "基础养老金（年）":                 "基础养老金",
    "社平工资（年）":                   "社平工资",
    "缴费指数（下限）":                 "缴费指数下限",
    "缴费指数（上限）":                 "缴费指数上限",
    "可选档次规则":                     "可选档次规则",
    "缴费基数（年，下限）":             "缴费基数下限",
    "缴费基数（年，上限）":             "缴费基数上限",
    "缴费金额（年，下限）":             "缴费金额下限",
    "缴费金额（年，上限）":             "缴费金额上限",
}

def get_policy(prov_name):
    p = policy_by_prov.get(prov_name)
    if p is None:
        return None
    out = {}
    for orig, renamed in COL_RENAME.items():
        v = p.get(orig)
        if renamed in ("缴费档次与补贴明细", "可选档次规则"):
            out[renamed] = str(v).strip() if v else "不清楚"
        else:
            try:
                out[renamed] = str(int(float(v))) if v is not None else "不清楚"
            except (ValueError, TypeError):
                out[renamed] = str(v).strip() if v else "不清楚"
    return out

# Individual-income terciles over the full candidate pool
log("\n[6] Computing individual-income terciles from filtered_flexible_workers...")
# Read I3a_6 for all candidate-pool respondents.
filtered_inc = ind_full[ind_full["_iid_f"].isin(filtered_iids_float)]["I3a_6"].copy()
valid_inc = filtered_inc[~filtered_inc.apply(is_missing)]
valid_inc = valid_inc.dropna().astype(float)
valid_inc = valid_inc[valid_inc >= 0]

q33 = valid_inc.quantile(1/3) if len(valid_inc) > 0 else 0
q67 = valid_inc.quantile(2/3) if len(valid_inc) > 0 else 0

log(f"  Valid income records: {len(valid_inc)}")
log(f"  Terciles: low <= {q33:.0f} CNY, middle <= {q67:.0f} CNY, high > {q67:.0f} CNY")

def income_group(v):
    if v is None or is_missing(v): return "不清楚"
    try:
        f = float(v)
        if math.isnan(f) or f < 0: return "不清楚"
        if f <= q33: return "低收入"
        if f <= q67: return "中收入"
        return "高收入"
    except (ValueError, TypeError):
        return "不清楚"

# Reconstruct sampled cases
log("\n[7] Reconstructing variables row by row...")

results = []
skip_policy = 0

for _, srow in sampled.iterrows():
    iid_f = srow["_iid_f"]
    fid_f = srow["_fid_f"]
    stratum = srow.get("stratum", "")

    # Match the individual table using float64 IDs.
    pmatch = ind[ind["_iid_f"] == iid_f]
    if len(pmatch) == 0:
        log(f"  Skipped: no individual matched IID_float={iid_f}")
        continue
    p = pmatch.iloc[0]

    # Match the household table using float64 IDs.
    f = fam.loc[fid_f] if fid_f in fam.index else None

    # Basic demographics
    gender_code = safe_val(p.get("Igender"))
    gender = {1: "男", 2: "女"}.get(gender_code, "不清楚")

    birthyear = safe_val(p.get("birthyear"))
    age = (SURVEY_YEAR - int(birthyear)) if birthyear and not is_missing(birthyear) else "不清楚"

    edu_code = safe_val(p.get("I2_1"))
    education = EDU_MAP.get(edu_code, "不清楚") if edu_code else "不清楚"

    health_code = safe_val(p.get("I9_4_1"))
    health = HEALTH_MAP.get(health_code, "不清楚") if health_code else "不清楚"

    # Hukou information
    hukou_code = safe_val(p.get("I1_14"))
    hukou_type = HUKOU_MAP.get(hukou_code, "不清楚") if hukou_code else "不清楚"

    # Hukou province: first two digits of the six-digit code.
    hukou_psu = str(safe_val(p.get("I1_3_1_psu"), "")).strip()
    if hukou_psu and len(hukou_psu) >= 2 and hukou_psu not in {"nan", ""}:
        try:
            hukou_prov_code = int(hukou_psu[:2])
        except ValueError:
            hukou_prov_code = None
    else:
        hukou_prov_code = None
    hukou_prov_name = PROV_CODE_TO_NAME.get(hukou_prov_code, "不清楚")

    # Province of residence
    res_prov_raw = safe_val(p.get("PROV2018"))
    try:
        res_prov_code = int(float(res_prov_raw)) if res_prov_raw else None
    except (ValueError, TypeError):
        res_prov_code = None
    res_prov_name = PROV_CODE_TO_NAME.get(res_prov_code, "不清楚")

    # Migration status
    if hukou_prov_code and res_prov_code:
        migration = "是" if hukou_prov_code != res_prov_code else "否"
    else:
        migration = "不清楚"

    # Urban/rural residence
    areatype_code = safe_val(p.get("areatype"))
    urban_rural = {1: "农村", 2: "城镇"}.get(areatype_code, "不清楚")

    # Region, based on province of residence
    region = REGION_MAP.get(res_prov_code, "不清楚")

    # Employment
    i3a9 = safe_val(p.get("I3a_9"))
    i3a16 = safe_val(p.get("I3a_16"))
    i3a1_5 = safe_val(p.get("I3a1_5"))

    # Employment arrangement
    if i3a9 in (9, 11, 12):
        work_nature = UNIT_MAP.get(i3a9, "不清楚")
    elif i3a9 in (1, 2, 3, 4, 5, 6, 7, 8) and i3a1_5 == 2:
        base = WORK_NATURE_MAP.get(i3a16, "灵活雇员")
        work_nature = f"{base}（无合同）"
    else:
        work_nature = WORK_NATURE_MAP.get(i3a16, "不清楚")

    # Industry
    ind_code = safe_val(p.get("I3a_8"))
    industry = IND_MAP.get(ind_code, "不清楚") if ind_code else "不清楚"

    # Occupation: first digit of the five-digit code.
    occ_raw = safe_val(p.get("I3a_7code"))
    if occ_raw and not is_missing(occ_raw):
        try:
            occ_first = int(str(int(occ_raw))[0])
            occupation = OCC_MAP.get(occ_first, "不清楚")
        except (ValueError, TypeError):
            occupation = "不清楚"
    else:
        occupation = "不清楚"

    # Employer type
    unit_type = UNIT_MAP.get(i3a9, "不清楚") if i3a9 else "不清楚"

    # Labor-contract status
    contract = {1: "已签合同", 2: "未签合同"}.get(i3a1_5, "不清楚")

    # Individual income
    inc_raw = safe_val(p.get("I3a_6"))
    if inc_raw is not None and not is_missing(inc_raw):
        try:
            inc_val = float(inc_raw)
            inc_str = fmt_yuan(inc_val)
        except (ValueError, TypeError):
            inc_val = None
            inc_str = "不清楚"
    else:
        inc_val = None
        inc_str = "不清楚"

    inc_group = income_group(inc_raw)

    # Household finances from the household table
    def fget(col, default=None):
        if f is None: return default
        v = f.get(col) if hasattr(f, 'get') else (f[col] if col in f.index else default)
        return v

    fam_inc = safe_float(fget("f4_1"),  default=float('nan'))   # Missing maps to unknown; a true zero remains 0 CNY.
    fam_exp = safe_float(fget("f4_25"), default=float('nan'))   # Missing maps to unknown; a true zero remains 0 CNY.
    fam_debt = safe_float(fget("f4_17_1"))
    fam_pension_year = safe_float(fget("f4_8"))
    fam_pension_month = fam_pension_year / 12 if fam_pension_year > 0 else 0.0

    # Household assets, summed across components
    # Residential property: convert 10,000 CNY to CNY.
    house_self_built = safe_float(fget("f3_4_4")) * 10000
    house_purchased  = safe_float(fget("f3_5_6")) * 10000
    house_shared     = safe_float(fget("f3_6_6")) * 10000
    # Other housing assets (CNY)
    other_houses = sum(
        safe_float(fget(f"f3_11_5_{i}"))
        for i in range(1, 5)
    )
    # Property in the home village (CNY)
    hometown_house = safe_float(fget("f5f5_2b"))
    # Vehicles (CNY)
    car = max(safe_float(fget("f4_10_3")), safe_float(fget("f4_10_3_panel")))
    moto       = safe_float(fget("f4_11_2_panel"))
    tractor    = safe_float(fget("f4_12_2_panel"))
    farm_equip = safe_float(fget("f4_13_2_panel"))
    livestock  = safe_float(fget("f4_14_2"))
    # Accounts receivable (CNY)
    receivable = safe_float(fget("f4_16_1"))
    # Business assets from the individual table: CNY plus 10,000 CNY converted to CNY.
    biz_employer = safe_float(p.get("I3a2_4_total"))              # CNY
    biz_self     = safe_float(p.get("I3a3_23")) * 10000           # Convert 10,000 CNY to CNY.

    total_asset = (
        house_self_built + house_purchased + house_shared +
        other_houses + hometown_house +
        car + moto + tractor + farm_equip + livestock +
        receivable + biz_employer + biz_self
    )

    # Household composition
    fam_size_raw = fget("number")
    try:
        fam_size = int(safe_float(fam_size_raw, 0))
    except (ValueError, TypeError):
        fam_size = 0

    children = 0
    elderly  = 0
    for i in range(1, 23):
        by_raw = fget(f"birthyear_{i}")
        if by_raw is None or is_missing(by_raw): continue
        try:
            by = int(float(by_raw))
        except (ValueError, TypeError):
            continue
        if by <= 0 or by > SURVEY_YEAR: continue
        member_age = SURVEY_YEAR - by
        if member_age < 16:
            children += 1
        elif member_age >= 60:
            elderly += 1

    # Household enrollment and benefit receipt
    fam_insured_cnt = int(srow["家庭参保人数"])
    fam_receiving   = "不清楚"   # CLDS does not provide this variable.

    # Psychological characteristics
    risk_code = safe_val(p.get("I7_14_6_w16"))
    risk_pref = RISK_MAP.get(risk_code, "不清楚") if risk_code and not is_missing(risk_code) else "不清楚"

    expect_code = safe_val(p.get("I7_6_3"))
    econ_expect = EXPECT_MAP.get(expect_code, "不清楚") if expect_code and not is_missing(expect_code) else "不清楚"

    # Enrollment history, unavailable in CLDS
    hist_type       = "不清楚"
    hist_years      = "不清楚"
    hist_gap        = "不清楚"

    # Policy matching
    policy_prov = hukou_prov_name
    policy = get_policy(policy_prov)
    if policy is None:
        # Fall back to the province of residence.
        policy = get_policy(res_prov_name)
    if policy is None:
        skip_policy += 1
        policy = {k: "不清楚" for k in COL_RENAME.values()}

    # Ground-truth enrollment decision
    ins_type = srow.get("ins_type", "不参保")
    if ins_type == "城镇职工养老保险":
        decision = "参保"
        account  = "城镇职工养老保险"
    elif ins_type == "城乡居民养老保险":
        decision = "参保"
        account  = "城乡居民养老保险"
    else:
        decision = "不参保"
        account  = "不参保"

    # Assemble the reconstructed row.
    row_out = {
        "家庭ID":      str(int(fid_f)) if not math.isnan(fid_f) else str(fid_f),
        "个人ID":      str(int(iid_f)) if not math.isnan(iid_f) else str(iid_f),
        "stratum":     stratum,
        "年龄":        age,
        "性别":        gender,
        "文化程度":    education,
        "健康状况":    health,
        "户口性质":    hukou_type,
        "户口省份":    hukou_prov_name,
        "常住省份":    res_prov_name,
        "是否流动":    migration,
        "城乡":        urban_rural,
        "地区":        region,
        "工作性质":    work_nature,
        "工作行业":    industry,
        "工作职业":    occupation,
        "单位类型":    unit_type,
        "是否签劳动合同": contract,
        "个人年收入":  inc_str,
        "收入分组":    inc_group,
        "家庭总收入":  fmt_yuan(fam_inc),
        "家庭总消费":  fmt_yuan(fam_exp),
        "家庭总资产":  fmt_yuan(total_asset),
        "家庭总负债":  fmt_yuan(fam_debt),
        "家庭人数":    fam_size,
        "子女数":      children,
        "老人数":      elderly,
        "历史参保类型": hist_type,
        "累计缴纳年限": hist_years,
        "是否存在断缴": hist_gap,
        "家庭参保人数": fam_insured_cnt,
        "家庭领取人数": fam_receiving,
        "家庭月均养老金": fmt_yuan(fam_pension_month),
        "风险偏好":    risk_pref,
        "经济预期":    econ_expect,
        "参保决策":    decision,
        "参保账户":    account,
        # Embed policy fields in the CSV for direct use by task 4.
        "缴费档次与补贴明细": policy["缴费档次与补贴明细"],
        "基础养老金":         policy["基础养老金"],
        "社平工资":           policy["社平工资"],
        "缴费指数下限":       policy["缴费指数下限"],
        "缴费指数上限":       policy["缴费指数上限"],
        "可选档次规则":       policy["可选档次规则"],
        "缴费基数下限":       policy["缴费基数下限"],
        "缴费基数上限":       policy["缴费基数上限"],
        "缴费金额下限":       policy["缴费金额下限"],
        "缴费金额上限":       policy["缴费金额上限"],
    }
    results.append(row_out)

log(f"\n  Reconstruction complete: {len(results)} records; policy unmatched: {skip_policy}")

out_df = pd.DataFrame(results)

# Data-quality report
log("\n[8] Data-quality report...")
log(f"\n  Enrollment decision distribution:")
for v, cnt in out_df["参保决策"].value_counts().items():
    log(f"    {v}: {cnt}")

log(f"\n  Enrollment account distribution:")
for v, cnt in out_df["参保账户"].value_counts().items():
    log(f"    {v}: {cnt}")

log(f"\n  Gender distribution: {out_df['性别'].value_counts().to_dict()}")
log(f"  Urban-rural distribution: {out_df['城乡'].value_counts().to_dict()}")
log(f"  Region distribution: {out_df['地区'].value_counts().to_dict()}")
log(f"  Income-group distribution: {out_df['收入分组'].value_counts().to_dict()}")

log(f"\n  Risk-preference distribution:")
for v, cnt in out_df["风险偏好"].value_counts().items():
    log(f"    {v}: {cnt}")

log(f"\n  Household enrollment count summary:")
fam_ins_num = pd.to_numeric(out_df["家庭参保人数"], errors="coerce")
log(f"    Mean: {fam_ins_num.mean():.2f}, median: {fam_ins_num.median():.0f}, maximum: {fam_ins_num.max():.0f}")

log(f"\n  Policy matching:")
log(f"    Records with 缴费档次与补贴明细='不清楚': {(out_df['缴费档次与补贴明细'] == '不清楚').sum()}")

# Save reconstructed cases
log(f"\n[9] Saving...")
out_df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
log(f"  Saved: {OUT_CSV}")
log(f"  Rows: {len(out_df)}  Columns: {len(out_df.columns)}")
log("\nDone.")
_log.close()
